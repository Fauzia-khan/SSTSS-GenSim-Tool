import os
import time
import subprocess
import threading
import signal

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QFormLayout,
    QMessageBox, QDialog, QScrollArea, QCheckBox, QLineEdit, QComboBox,
    QTableWidget, QTableWidgetItem, QFileDialog
)
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, SubElement, ElementTree
from PyQt5.QtGui import QPixmap, QFont
from PyQt5.QtCore import Qt


class ViewInformationWindow(QDialog):
    def __init__(self, excel_filename=None, scenario_row=None):
        super().__init__()

        scenario_row += 1  # skip header row

        self.setWindowTitle('Scenario Parameter Configuration')
        self.setGeometry(300, 200, 500, 300)
        self.global_light_condition = None
        wb = load_workbook(excel_filename)
        ws = wb.active
        self.setStyleSheet("background-color: #e9f0fa;")   # Azure-like blue

        # Main layout
        main_layout = QHBoxLayout()

        # Left side: tags
        tag_layout = QFormLayout()
        tag_layout.addRow(QLabel("<b>Tags</b>"), QLabel(""))

        tag_data = {}

        # Column ranges
        actor_column_start, actor_column_end = 5, 10
        weather_column_start, weather_column_end = 11, 14
        light_column_start, light_column_end = 15, 17
        behaviour_column_start, behaviour_column_end = 18, 32
        road_topology_column_start, road_topology_column_end = 33, 35
        scenario_group_column = 38

        # Actors
        for i in range(actor_column_start, actor_column_end):
            if ws.cell(row=scenario_row, column=i).value == 1:
                tag_data['Actors'] = [ws.cell(row=1, column=i).value.split('_')[1]]

        # Weather
        for i in range(weather_column_start, weather_column_end):
            if ws.cell(row=scenario_row, column=i).value == 1:
                tag_data['Weather'] = [ws.cell(row=1, column=i).value.split('_')[1]]

        # Light
        for i in range(light_column_start, light_column_end):
            if ws.cell(row=scenario_row, column=i).value == 1:
                tag_data['Light'] = [ws.cell(row=1, column=i).value.split('_')[1]]
                self.global_light_condition = tag_data['Light']

        # Driving Maneuver
        for i in range(behaviour_column_start, behaviour_column_end):
            if ws.cell(row=scenario_row, column=i).value == 1:
                tag_data['Driving Maneuver'] = [ws.cell(row=1, column=i).value.split('_')[1]]

        # Road Topology
        for i in range(road_topology_column_start, road_topology_column_end):
            if ws.cell(row=scenario_row, column=i).value == 1:
                tag_data['Road Topology'] = [ws.cell(row=1, column=i).value.split('_')[1]]

        # Scenario Group
        scenario_group = ws.cell(row=scenario_row, column=scenario_group_column).value
        if scenario_group:
            tag_data['Scenario Group'] = [scenario_group]

        for category, items in tag_data.items():
            tag_layout.addRow(QLabel(f"<b>{category}</b>"), QLabel(""))
            for item in items:
                tag_layout.addRow(QLabel(f"    {item}"), QLabel(""))

        # Right side: controls
        files_layout = QFormLayout()
        files_layout.addRow(QLabel("<b>Files</b>"), QLabel(""))

        # Inputs
        self.speed_input = QLineEdit()
        self.speed_input.setPlaceholderText("Enter ego vehicle speed")
        self.other_actor_speed_input = QLineEdit()
        self.other_actor_speed_input.setPlaceholderText("Enter other actor speed")
        self.other_actor_distance_input = QLineEdit()
        self.other_actor_distance_input.setPlaceholderText("Enter other actor distance")
        self.timeout_input = QLineEdit()
        self.timeout_input.setPlaceholderText("Enter simulation duration")

        # Ego vehicle dropdown
        self.ego_vehicle_dropdown = QComboBox()
        vehicle_options = ["Nissan.patrol","Tesla Model3", "Tesla Cybertruck", "Audi A2", "BMW X5", "Mercedes C-Class"]
        self.ego_vehicle_dropdown.addItems(vehicle_options)
        files_layout.addRow(QLabel("Select Model for other vehicle"), self.ego_vehicle_dropdown)

        # Map dropdown
        self.map_dropdown = QComboBox()
        map_options = ["Town01", "Town02", "Town03", "Town04", "Town05", "Town06", "Town07"]
        self.map_dropdown.addItems(map_options)
        files_layout.addRow(QLabel("Select Map"), self.map_dropdown)

        # Buttons
        execute_sim_button = QPushButton("Execute Simulation")
        execute_sim_button.setFixedSize(150, 25)
        execute_sim_button.clicked.connect(self.execute_simulation)

        self.show_results_button = QPushButton("Show Results")
        self.show_results_button.setFixedSize(150, 25)
        self.show_results_button.setEnabled(False)
        self.show_results_button.clicked.connect(self.show_results)

        # Add inputs
        files_layout.addRow(QLabel("Select Speed for Ego vehicle"), self.speed_input)
        files_layout.addRow(QLabel("Enter Speed of Other Actor"), self.other_actor_speed_input)
        files_layout.addRow(QLabel("Enter Distance of Other Actor"), self.other_actor_distance_input)
        files_layout.addRow(QLabel("Enter Simulation Duration"), self.timeout_input)
        files_layout.addRow(execute_sim_button)
        files_layout.addRow(self.show_results_button)

        # Add layouts to main
        main_layout.addLayout(tag_layout)
        main_layout.addLayout(files_layout)
        self.setLayout(main_layout)

    # ------------------- Simulation Control -------------------

    def run_scenario_runner(self, summary_log):
        scenario_runner_script = "/home/laima/Desktop/SSTSS Tool 1st september/InterfaceProject/run_scenario_runner.sh"
        print(f"[INFO] Running ScenarioRunner from {scenario_runner_script}")
        print(f"[INFO] Saving output to {summary_log}")

        with open(summary_log, "w", encoding="utf-8", errors="ignore") as f:
            process = subprocess.Popen(
                ["bash", scenario_runner_script],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True
            )
            for line in process.stdout:
                print(line, end="")
                f.write(line)
            process.wait()

        print(f"[✓] ScenarioRunner finished. Summary saved to {summary_log}")

    def stop_autoware(self):
        """Find and stop Autoware process by name."""
        try:
            result = subprocess.run(
                ["pgrep", "-f", "roslaunch autoware_mini start_carla.launch"],
                capture_output=True,
                text=True
            )
            pids = result.stdout.strip().splitlines()
            for pid in pids:
                print(f"[INFO] Stopping Autoware process {pid}")
                os.kill(int(pid), signal.SIGTERM)
            if not pids:
                print("[INFO] No Autoware process found to stop.")
        except Exception as e:
            print(f"[ERROR] Failed to stop Autoware: {e}")

    def execute_simulation(self):
        print("In: Execute Simulation")
        ego_vehicle_speed = self.speed_input.text()
        other_vehicle_distance = self.other_actor_distance_input.text()
        other_vehicle_speed = self.other_actor_speed_input.text()
        timeout = self.timeout_input.text()

        # Update Scenario Template
        with open('scenario_files/python_files/template_follow_leading_vehicle.py', 'r') as f:
            lines = f.readlines()
        lines[7] = f'timeout = {timeout}\n'
        lines[8] = f'other_vehicle_distance = {other_vehicle_distance}\n'
        lines[9] = f'other_vehicle_speed = {other_vehicle_speed}\n'
        with open(
            '/home/laima/Documents/scenario_runner-master/srunner/scenarios/follow_leading_vehicle.py',
            'w', encoding='utf-8'
        ) as f:
            f.writelines(lines)

        # Start CARLA
        subprocess.Popen(["gnome-terminal", "--", "bash", "-c", "./generate_simulation.sh; exec bash"])
        time.sleep(10)

        # Start Autoware
        cmd = (
            "source /opt/ros/noetic/setup.bash && "
            "source ~/ali/ali_ws/devel/setup.bash && "
            "source ~/Documents/autoware_mini_ws/devel/setup.bash && "
            "export CARLA_ROOT=$HOME/Documents/CARLA_0.9.13 && "
            "export PYTHONPATH=/home/laima/ali/ali_ws/devel/lib/python3/dist-packages:"
            "/home/laima/Documents/autoware_mini_ws/devel/lib/python3/dist-packages:"
            "/opt/ros/noetic/lib/python3/dist-packages:"
            "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla/dist/carla-0.9.13-py3.7-linux-x86_64.egg:"
            "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla/agents:"
            "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla && "
            f"roslaunch autoware_mini start_carla.launch map_name:=Town01 generate_traffic:=false speed_limit:=10; exec bash"
        )
        subprocess.Popen(["terminator", "-e", f"bash -c '{cmd}'"])
        time.sleep(7)

        # Run ScenarioRunner
        results_dir = "/home/laima/Documents/scenario_runner-master/results/test"
        os.makedirs(results_dir, exist_ok=True)
        summary_log = os.path.join(results_dir, "scenario_summary.log")
        runner_thread = threading.Thread(target=self.run_scenario_runner, args=(summary_log,))
        runner_thread.start()

        # Run set_goal
        os.system('gnome-terminal -- bash -c "./run_set_goal.sh; exec bash"')

        # Wait for ScenarioRunner
        runner_thread.join()

        # Stop Autoware
        self.stop_autoware()

        # Run Metrics
        try:
            subprocess.run(
                [
                    "python3.8", "metrics_manager.py",
                    "--metric", "srunner/metrics/examples/velocity_and_distance_metric.py",
                    "--log", "results/test/FollowLeadingVehicle_1.log"
                ],
                cwd="/home/laima/Documents/scenario_runner-master",
                env={
                    **os.environ,
                    "CARLA_ROOT": "/home/laima/Documents/CARLA_0.9.13",
                    "PYTHONPATH": (
                        "/home/laima/ali/ali_ws/devel/lib/python3/dist-packages:"
                        "/home/laima/Documents/autoware_mini_ws/devel/lib/python3/dist-packages:"
                        "/opt/ros/noetic/lib/python3/dist-packages:"
                        "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla/dist/carla-0.9.13-py3.7-linux-x86_64.egg:"
                        "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla/agents:"
                        "/home/laima/Documents/CARLA_0.9.13/PythonAPI/carla"
                    )
                },
                check=True
            )
            print("[✓] Metrics script finished.")
            self.show_results_button.setEnabled(True)
        except subprocess.CalledProcessError as e:
            print(f"[ERROR] Failed to run metrics script: {e}")
            return

    # ------------------- Results Display -------------------

    def show_results(self):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QScrollArea, QWidget, QPushButton, QFileDialog
        import re, zipfile

        results_dir = "/home/laima/Documents/scenario_runner-master/results/test"
        if not os.path.exists(results_dir):
            QMessageBox.warning(self, "Results Not Found", f"⚠️ Path does not exist:\n{results_dir}")
            return

        all_files = os.listdir(results_dir)
        pattern = re.compile(r"FollowScenario_(\d{8}_\d{6})")
        timestamps = [m.group(1) for f in all_files if (m := pattern.search(f))]
        latest_ts = max(timestamps) if timestamps else None

        selected_files = []
        if latest_ts:
            for f in all_files:
                if latest_ts in f and f.lower().endswith((".png", ".jpg", ".jpeg", ".csv")):
                    selected_files.append(os.path.join(results_dir, f))

        log_path = os.path.join(results_dir, "FollowLeadingVehicle_1.log")
        json_path = os.path.join(results_dir, "FollowLeadingVehicle_1.json")
        summary_path = os.path.join(results_dir, "scenario_summary.log")
        if os.path.exists(log_path): selected_files.append(log_path)
        if os.path.exists(json_path): selected_files.append(json_path)

        if not selected_files and not os.path.exists(summary_path):
            QMessageBox.warning(self, "No Results", "⚠️ No results found for latest simulation")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Simulation Report")
        dialog.resize(1000, 800)

        layout = QVBoxLayout()
        container = QWidget()
        container_layout = QVBoxLayout(container)

        for f in selected_files:
            if f.lower().endswith((".png", ".jpg", ".jpeg")):
                pixmap = QPixmap(f)
                if not pixmap.isNull():
                    img_label = QLabel()
                    img_label.setPixmap(pixmap.scaledToWidth(900, Qt.SmoothTransformation))
                    img_label.setAlignment(Qt.AlignCenter)
                    container_layout.addWidget(img_label)

        if os.path.exists(summary_path):
            with open(summary_path, "r", encoding="utf-8", errors="ignore") as f:
                summary_text = f.read()
            summary_label = QLabel(summary_text)
            summary_label.setFont(QFont("Courier", 10))
            summary_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            summary_label.setWordWrap(True)
            container_layout.addWidget(QLabel("---- Scenario Summary ----"))
            container_layout.addWidget(summary_label)

        def download_all():
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Results As",
                f"simulation_results_{latest_ts}.zip" if latest_ts else "simulation_results.zip",
                "Zip Files (*.zip)"
            )
            if save_path:
                with zipfile.ZipFile(save_path, 'w') as zipf:
                    for file_path in selected_files:
                        zipf.write(file_path, os.path.basename(file_path))
                    if os.path.exists(summary_path):
                        zipf.write(summary_path, os.path.basename(summary_path))
                print(f"[✓] All files saved to {save_path}")

        btn = QPushButton("Download Files")
        btn.clicked.connect(download_all)
        container_layout.addWidget(btn, alignment=Qt.AlignCenter)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(container)
        layout.addWidget(scroll)
        dialog.setLayout(layout)
        dialog.exec_()


    def backup_simulation_outputs(self, timestamp):
        import shutil
        import os
        import glob

        base_dir = "/home/laima/Documents/scenario_runner-master/results/test"
        backup_dir = os.path.join(base_dir, f"FollowLeadingVehicle_1_{timestamp}")
        os.makedirs(backup_dir, exist_ok=True)

        print(f"[INFO] Backing up outputs to: {backup_dir}")

        # Copy .log and .json
        for ext in ["log", "json"]:
            src = os.path.join(base_dir, f"FollowLeadingVehicle_1.{ext}")
            if os.path.exists(src):
                shutil.copy(src, backup_dir)
                print(f"[✓] Copied {src}")
            else:
                print(f"[!] Missing: {src}")

        # Copy metrics files with timestamp
        for ext in ["metrics.csv", "speed_distance.png", "jerk_1hz.png"]:
            pattern = f"FollowScenario_{timestamp}_{ext}"
            matches = glob.glob(os.path.join(base_dir, pattern))
            for f in matches:
                shutil.copy(f, backup_dir)
                print(f"[✓] Copied {f}")
            if not matches:
                print(f"[!] No match for: {pattern}")

    def adjust_weather_condition(self):
        file_path = '/home/laima/Documents/scenario_runner-master/srunner/examples/FollowLeadingVehicle.xml'

        # Check Light
        if self.global_light_condition[0].lower() == 'day':
            sun_altitude_angle = 90
        else:
            sun_altitude_angle = -10

        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        new_lines = []
        for line in lines:
            if "<weather" in line:
                import re
                # replace cloudiness
                line = re.sub(r'cloudiness="[^"]+"', 'cloudiness="0"', line)
                # replace precipitation
                line = re.sub(r'precipitation="[^"]+"', 'precipitation="0"', line)

                #replace daylight
                line = re.sub(r'sun_altitude_angle="[^"]+"', f'sun_altitude_angle="{sun_altitude_angle}"', line)
                print("Daylight status set to", sun_altitude_angle)
            new_lines.append(line)

        with open(file_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)

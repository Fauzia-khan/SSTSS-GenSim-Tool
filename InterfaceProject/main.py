"""
Main interface script.
"""
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from PyQt5.QtWidgets import QApplication, QWidget,  QProgressDialog,QVBoxLayout, QPushButton, QComboBox, QLabel, QHBoxLayout, QGridLayout,QFrame
from PyQt5.QtWidgets import QLabel, QComboBox, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox
from PyQt5.QtCore import Qt, QTimer
from openpyxl.reader.excel import load_workbook

from add_scenario import AddScenarioWindow
from constants import scenarios_excel_file_name
from view_scenario_window import ViewScenariosWindow
from view_all_scenario_window import ViewAllScenariosWindow
from select_odd_window import SelectODDWindow
from select_scenarios_basedOn_ODD import SelectScenariosBasedOnOdd
from scenario_grouping import formulate_scenario_groups
from prioritize_scenario_group import prioritize_scenario_groups
from simulator_filtering import filter_scenarios_based_on_simulator
from select_scenario_window import SelectScenarioWindow
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import os

    ##################### Main Application Window
class ScenarioDatabaseApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('SSTSS GenSim Tool')
        self.setGeometry(300, 100, 600, 300)
        self.setFixedSize(500, 400)
        # ---------- TOOL BUTTON STYLE ----------
        self.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f9f9f9, stop:1 #e0e0e0
                );
                border: 1px solid #8c8c8c;     /* Thin, soft gray border */
                border-radius: 4px;            /* Light rounding for professional look */
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 10.5pt;
                color: #222;
                padding: 5px 10px;
                min-width: 100px;
            }

            QPushButton:hover {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 #fdfdfd, stop:1 #dcdcdc
                );
                border: 1px solid #666;
            }

            QPushButton:pressed {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 #dcdcdc, stop:1 #c0c0c0
                );
                border: 1px solid #555;
            }

            QPushButton:disabled {
                background-color: #f5f5f5;
                color: #999;
                border: 1px solid #ccc;
            }
        """)
        self.overlay_frame = QFrame(self)
        self.overlay_frame.setGeometry(70, 40, 370, 300)  # x, y, width, height
        self.overlay_frame.setStyleSheet("""
            QFrame {
                background-color: rgba(255, 255, 255, 160);  /* soft translucent white */
                border-radius: 0px;                          /* sharp corners */
                border: 1px solid rgba(255, 255, 255, 180);  /* subtle white outline */
            }
        """)

        # Background image
        self.background_label = QLabel(self)
        pixmap = QPixmap("Self.jpeg")
        self.background_label.setPixmap(pixmap)
        self.background_label.setScaledContents(True)
        self.background_label.setGeometry(self.rect())
        self.background_label.lower()

        # ✅ Create one main layout
        main_layout = QVBoxLayout(self)

        # ---------- TOP LAYOUT ----------

        up_layout = QHBoxLayout()

        # ---------- TITLE ----------
        self.title_label = QLabel("Scenario Selection", self)
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setFont(QFont("Arial", 20, QFont.Bold))
        self.title_label.setStyleSheet("color: white;")
        self.title_label.setFixedWidth(250)
        self.title_label.move(145, 50)

        # ---------- BUTTONS ----------
        self.view_all_button = self.create_button('1. Choose Catalog', 100, self.view_all_scenarios)
        self.odd_button = self.create_button('2. Select ODD', 150, self.open_select_odd)
        self.prioritize_scenario_button = self.create_button('3. Select Dataset', 200,
                                                             self.prioritize_scenario_groups_function)
        self.filter_scenario_button = self.create_button('4. Select Simulator', 250,
                                                         self.open_simulator_selection_window)

        # ---------- BUTTON CREATION FUNCTION ----------

    def create_button(self, text, y_pos, click_action):
        button = QPushButton(text, self)
        button.setFixedSize(200, 40)
        button.move(160, y_pos)
        button.setStyleSheet("""
                   QPushButton {
                       background-color: #ffffff;
                       border: 0px solid #555555;
                       border-radius: 6px;
                       font: bold 12pt 'Arial';
                       color: #222;
                       padding: 5px 10px;
                   }
                   QPushButton:hover {
                       background-color: #f2f2f2;
                       border: 2px solid #000000;
                   }
               """)
        button.clicked.connect(click_action)
        return button


    def ADD_ITEMS_TO_CATALOG_COMBO(self):

        print(os.getcwd()+"/"+scenarios_excel_file_name)
        wb = load_workbook(os.getcwd()+"/"+scenarios_excel_file_name)
        ws = wb.active
        print(23)
        catalogs = []
        for i in range(3, ws.max_row+1):
            catalogs.append(ws.cell(row=i, column=37).value)

        catalogs.append('US')
        catalogs.append('Singapore')
        catalogs.append('Europe')
        catalogs.append('Other')

        catalogs = list(set(catalogs))

        for c in catalogs:
            self.catalog_combo.addItem(c)

    def add_scenario(self):
        catalog = self.catalog_combo.currentText()
        self.add_window = AddScenarioWindow(catalog)
        self.add_window.show()

    def view_scenarios(self):
        catalog = self.catalog_combo.currentText()
        self.view_window = ViewScenariosWindow(catalog)
        self.view_window.exec_()

    def view_all_scenarios(self):
        self.view_all_window = ViewAllScenariosWindow()
        self.view_all_window.exec_()

    def remove_duplicates_function(self):
        QMessageBox.information(self, "Info", "Remove duplicates functionality will be added soon.")

    def open_select_odd(self):
        #self.odd_window = SelectODDWindow()
        #self.odd_window.exec_()
        self.odd_window = SelectODDWindow()
        self.odd_window.odd_saved.connect(self.select_scenarios_based_on_odd_function)
        self.odd_window.exec_()
    def select_scenarios_based_on_odd_function(self):
        self.select_scenario_odd = SelectScenariosBasedOnOdd()

    def formulate_scenario_groups_function(self):
        formulate_scenario_groups()

    def prioritize_scenario_groups_function(self):
        #prioritize_scenario_groups()
        self.dataset_window = QWidget()
        self.dataset_window.setWindowTitle("Select Dataset")
        self.dataset_window.setFixedSize(300, 200)
        self.dataset_window.setStyleSheet("""
                QWidget {
                    background-color: #e9f0fa;          /* Soft blue background */
                    font-family: 'Segoe UI', Arial, sans-serif;
                    font-size: 10.5pt;
                    color: #222;                         /* Dark gray text */
                }
                QPushButton {
                    background-color: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 #f9f9f9, stop:1 #e0e0e0
                    );
                    border: 1px solid #8c8c8c;
                    border-radius: 4px;
                    padding: 4px 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #f0f0f0;
                }
                QPushButton:pressed {
                    background-color: #dcdcdc;
                }
                 QLabel {
        font-weight: bold;
    }
            """)

        # Create layout and widgets
        layout = QVBoxLayout()

        # Add label
        label = QLabel("Select Dataset:", self.dataset_window)

        layout.addWidget(label)

        # Add dropdown menu (QComboBox) for dataset selection
        self.dataset_combo = QComboBox(self.dataset_window)
        self.dataset_combo.addItem("US")
        self.dataset_combo.addItem("EU")
        layout.addWidget(self.dataset_combo)


        # Add OK button
        ok_button = QPushButton("OK", self.dataset_window)
        ok_button.clicked.connect(self.apply_dataset_selection)
        layout.addWidget(ok_button)

        # Set layout for the dataset window
        self.dataset_window.setLayout(layout)

        # Show the dataset window
        self.dataset_window.show()

    # Function to apply the selected dataset
    def apply_dataset_selection(self):
        selected_dataset = self.dataset_combo.currentText()  # Get selected dataset from the dropdown
        self.overall_selected_dataset = selected_dataset

        try:
            # Prioritize scenario groups based on the selected dataset (US or EU)
            prioritize_scenario_groups(dataset=selected_dataset)


        except Exception as e:
            # Display error message in case of failure
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

        # Close the dataset window after applying the selection
        self.dataset_window.close()

    def open_scenario_selector(self):
        self.select_scenario_window = QWidget()
        #self.select_scenario_window.exec_()

        self.select_scenario_window.setWindowTitle("Select Dataset")
        self.select_scenario_window.setFixedSize(300, 200)

        layout = QVBoxLayout()
        label = QLabel("Select Dataset")
        layout.addWidget(label)


        self.scenario_combo = QComboBox(self.select_scenario_window)
        self.scenario_combo.addItem("US")
        self.scenario_combo.addItem("EU")
        self.scenario_combo.addItem("Other")
        #self.scenario_combo.move(600, 500)
        layout.addWidget(self.scenario_combo)

        # Button to confirm selection
        select_button = QPushButton("OK", self.select_scenario_window)
        select_button.clicked.connect(self.run_select_scenarios)
        layout.addWidget(select_button)

        self.select_scenario_window.setLayout(layout)
        self.select_scenario_window.show()

    def open_simulator_selection_window(self):
        # Open a new window to select the simulator
        self.simulator_window = QWidget()
        self.simulator_window.setWindowTitle("Select Simulator")
        self.simulator_window.setFixedSize(300, 200)
        self.simulator_window.setStyleSheet("""
               QWidget {
                   background-color: #e9f0fa;          /* Light professional blue */
                   font-family: 'Segoe UI', Arial, sans-serif;
                   font-size: 10.5pt;
                   color: #222;
               }
               QPushButton {
                   background-color: qlineargradient(
                       x1:0, y1:0, x2:0, y2:1,
                       stop:0 #f9f9f9, stop:1 #e0e0e0
                   );
                   border: 1px solid #8c8c8c;
                   border-radius: 4px;
                   padding: 4px 8px;
                   font-weight: bold;
               }
               QPushButton:hover {
                   background-color: #f0f0f0;
               }
               QPushButton:pressed {
                   background-color: #dcdcdc;
               }
               QComboBox {
                   background-color: #ffffff;
                   border: 1px solid #8c8c8c;
                   border-radius: 3px;
                   padding: 2px 6px;
               }
               QLabel {
                   font-weight: bold;
               }
           """)
        layout = QVBoxLayout()
        label = QLabel("Select Simulator")
        layout.addWidget(label)

        # ComboBox for selecting the simulator (Carla, Gazebo, Audacity, LGSVL)
        self.simulator_combo = QComboBox(self.simulator_window)
        self.simulator_combo.addItem("Carla")
        self.simulator_combo.addItem("Gazebo")
        self.simulator_combo.addItem("Audacity")
        self.simulator_combo.addItem("LGSVL")
        layout.addWidget(self.simulator_combo)

        # Button to confirm selection
        select_button = QPushButton("OK", self.simulator_window)
        #select_button.clicked.connect(self.run_filter_scenarios)
        select_button.clicked.connect(self._handle_simulator_selection)
        layout.addWidget(select_button)

        self.simulator_window.setLayout(layout)
        self.simulator_window.show()

    def _handle_simulator_selection(self):
        """Handle the simulator selection and show progress dialog before running scenarios."""
        selected_simulator = self.simulator_combo.currentText()

        # Optional: store selected simulator if you need it elsewhere
        self.selected_simulator = selected_simulator

        # Close the simulator window
        self.simulator_window.close()

        # Show a progress/loading dialog
        progress = QProgressDialog("Selecting scenarios in progress...", None, 0, 0, self)
        progress.setWindowModality(Qt.ApplicationModal)
        progress.setWindowTitle("Please Wait")
        progress.setCancelButton(None)
        progress.setMinimumDuration(0)
        progress.show()

        # Use a short delay so the progress window appears before running heavy code
        QTimer.singleShot(600, lambda: self._run_scenarios_with_loading(progress))

    def _run_scenarios_with_loading(self, progress_dialog):
        """Run scenario selection after simulator selection."""
        try:
            self.run_select_scenarios()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
        finally:
            progress_dialog.close()




    def run_select_scenarios(self):
        # Use the dataset selected earlier in step 3
        if not hasattr(self, "overall_selected_dataset"):
            QMessageBox.warning(self, "Dataset Not Selected",
                                "Please select a dataset first (Step 3: Select Dataset).")
            return

        selected_database = self.overall_selected_dataset.lower()

        try:
            # Run dataset-specific selector
            if selected_database == 'us':
                from selected_scenarios import run_us_based_scenario_selector
                run_us_based_scenario_selector()

            elif selected_database == 'eu':
                from selected_scenarios import run_eu_based_scenario_selector
                run_eu_based_scenario_selector()

            else:
                QMessageBox.warning(self, "Unsupported Dataset",
                                    f"No selector available for: {self.overall_selected_dataset}")
                return

            # Open the SelectScenarioWindow
            from select_scenario_window import SelectScenarioWindow
            self.ssw_window = SelectScenarioWindow(self.overall_selected_dataset)
            self.ssw_window.exec_()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

    def run_filter_scenarios(self):
        # Based on the simulator selection, run the filter function
        selected_simulator = self.simulator_combo.currentText()
        selected_database = self.overall_selected_dataset

        if selected_simulator == "Carla":
            # Call the filter_scenarios_based_on_simulator function from an external file
            from simulator_filtering import filter_scenarios_based_on_simulator
            filter_scenarios_based_on_simulator('Carla', selected_database)
        elif selected_simulator == "Gazebo":
            from simulator_filtering import filter_scenarios_based_on_simulator
            filter_scenarios_based_on_simulator('Gazebo', selected_database)
        elif selected_simulator == "Audacity":
            from simulator_filtering import filter_scenarios_based_on_simulator
            filter_scenarios_based_on_simulator('Audacity', selected_database)
        elif selected_simulator == "LGSVL":
            from simulator_filtering import filter_scenarios_based_on_simulator
            filter_scenarios_based_on_simulator('LGSVL', selected_database)

        # Close the simulator window after selection
        self.simulator_window.close()

    def update_flag(self, index):
        # Get the actual catalog name text
        catalog_name = self.catalog_combo.currentText().lower()

        # Match catalog text to flag file
        flag_map = {
            "us": "usa_flag.png",
            "singapore": "singa_flag.png",
            "europe": "eu_flag.png",
            "other": "world_flag.png"
        }

        # Default to 'world_flag.png' if not found
        flag_path = flag_map.get(catalog_name, "world_flag.png")

        pixmap = QPixmap(flag_path)
        self.flag_label.setPixmap(pixmap)


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    window = ScenarioDatabaseApp()
    window.show()
    sys.exit(app.exec_())



#
        # # FORMULATE SCENARIO GROUPS BUTTON
        # self.formulate_scenario_button = QPushButton('4. Assign Scenarios to Scenario Groups', self)
        # self.formulate_scenario_button.setFixedSize(300, 45)
        # self.formulate_scenario_button.clicked.connect(self.formulate_scenario_groups_function)
        # self.formulate_scenario_button.move(400, 350)
        # #down_layout.addWidget(self.formulate_scenario_button)
    # FORMULATE SCENARIO GROUPS BUTTON
    ''' self.formulate_scenario_button = QPushButton('2. Assign Scenarios to Scenario Groups', self)
    self.formulate_scenario_button.setFixedSize(300, 45)
    self.formulate_scenario_button.clicked.connect(self.formulate_scenario_groups_function)
    self.formulate_scenario_button.move(400, 250)

    self.remove_duplicates_button = QPushButton('3. Remove Duplicates', self)
    self.remove_duplicates_button.setFixedSize(240, 45)
    self.remove_duplicates_button.clicked.connect(self.remove_duplicates_function)
    self.remove_duplicates_button.move(430, 300)
    down_layout = QHBoxLayout()
    self.select_scenario_odd_button = QPushButton('5. Select Scenarios Based On ODD', self)
    self.select_scenario_odd_button.setFixedSize(300, 45)
    self.select_scenario_odd_button.clicked.connect(self.select_scenarios_based_on_odd_function)
    self.select_scenario_odd_button.move(400, 400)
    main_layout.addLayout(down_layout)

    self.setLayout(main_layout)
    
    
    '''
'''
        self.selected_scenario_button = QPushButton('8. Selected Scenarios', self)
        self.selected_scenario_button.setFixedSize(180, 40)
        self.selected_scenario_button.move(150, 300)
        #self.selected_scenario_button.clicked.connect(self.open_scenario_selector)
        self.selected_scenario_button.clicked.connect(self.run_select_scenarios)

        self.add_button = QPushButton('Add Scenario', self)
        self.add_button.setFixedSize(140, 40)
        self.add_button.move(140, 330)
        self.add_button.clicked.connect(self.add_scenario)
        '''

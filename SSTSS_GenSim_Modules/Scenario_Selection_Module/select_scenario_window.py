import os

from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QLabel, QFormLayout, QMessageBox, QDialog, QScrollArea, QCheckBox
from modules.core.utils import delete_scenario_from_excel_file
from modules.constants import scenarios_excel_file_name
from openpyxl import load_workbook
from Scenario_Configuration_Module.scenario_parameter_configuration_window import ViewInformationWindow

from functools import partial



class SelectScenarioWindow(QDialog):
    def __init__(self, catalog_type=None):
        super().__init__()
        self.setWindowTitle("List of Selected Scenarios")
        self.setGeometry(300, 200, 500, 300)
        self.setStyleSheet("background-color: #e9f0fa;")
        self.setSizeGripEnabled(True)
        self.setMinimumSize(500, 300)


        # Scrollable Area
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        content_widget = QWidget()
        scroll_layout = QVBoxLayout(content_widget)

        self.checkboxes = []  # Store checkbox references for deletion

        scenarios = []



        print('+++++++++++++++++')
        #print('catalog_type', catalog_type)


        if catalog_type.lower() == 'us':
           excel_filename = 'selected_scenarios_us.xlsx'
        if catalog_type.lower() == 'eu':
           excel_filename = 'selected_scenarios_eu.xlsx'

        self.excel_filename = excel_filename

        wb = load_workbook(excel_filename) ## EXCEL FILE
        ws = wb.active

        #print('checkpoint X')
        for i in range(2, ws.max_row + 1):
            _temp_scenario_id = ws.cell(row=i, column=2).value
            each_scenario = [ws.cell(row=i, column=2).value, #scenario_id
                             ws.cell(row=i, column=3).value, #scenario_name
                             ws.cell(row=i, column=4).value, #scenario_description
                             ws.cell(row=i, column=36).value, #image_path
                             ws.cell(row=i, column=38).value, #scenario_group
                             ws.cell(row=i, column=44).value,  # priority
                             ]
            #print('checkpoint1', ws.cell(row=i, column=36).value)
            scenarios.append(each_scenario)


        scenario_view_counter = 0
        if scenarios:
            for scenario in scenarios:
                scenario_id, name, description, image_path, scenario_group, priority = scenario[0], scenario[1], scenario[2], scenario[3],scenario[4],scenario[5]

                form_layout = QFormLayout()

                scenario_view_counter += 1
                _view_button = QPushButton('View', self)
                _view_button.setFixedSize(50, 40)
                #_view_button.clicked.connect(self.view_button_callback)
                _view_button.clicked.connect(partial(self.view_button_callback, scenario_view_counter))

                checkbox = QCheckBox()
                self.checkboxes.append((checkbox, scenario_id))  # Store checkbox and associated scenario_id
                form_layout.addRow(checkbox)

                form_layout.addRow("ID:", QLabel(scenario_id))
                form_layout.addRow("Name:", QLabel(name))
                form_layout.addRow("Description:", QLabel(description))
                form_layout.addRow("ScenarioGroup:", QLabel(scenario_group))
                form_layout.addRow("Priority:", QLabel(str(priority)))

                if image_path:
                    #print('checkpoint 2', image_path)
                    image_item = QLabel()
                    pixmap = QPixmap(os.getcwd() + '/images/' + image_path).scaled(100, 100)
                    image_item.setPixmap(pixmap)
                    form_layout.addRow("Image:", image_item)
                else:
                    form_layout.addRow("Image:", QLabel("No Image"))

                form_layout.addRow('', _view_button)
                form_layout.addRow('---------------------------', QLabel('------------------------------'))
                scroll_layout.addLayout(form_layout)

            scroll_area.setWidget(content_widget)


    def view_button_callback(self, scenario_row):

        ## OPEN A WINDOW

        print("View for Scenario:", scenario_row)

        self.view_information_window = ViewInformationWindow(excel_filename=self.excel_filename,
                                                             scenario_row=scenario_row)
        self.view_information_window.exec_()























"""

            #layout.addWidget(scroll_area)

            # Delete Selected Button
        #     self.delete_button = QPushButton("Delete Selected Scenarios")
        #     self.delete_button.clicked.connect(self.delete_selected_scenarios)
        #     layout.addWidget(self.delete_button)
        #
        # else:
        #     layout.addWidget(QLabel("No scenarios available."))
        #
        # self.setLayout(layout)

    # def delete_selected_scenarios(self):
    #     #Delete selected scenarios from the database.
    #     selected_ids = [scenario_id for checkbox, scenario_id in self.checkboxes if checkbox.isChecked()]
    #
    #     if not selected_ids:
    #         QMessageBox.warning(self, "No Selection", "Please select at least one scenario to delete.")
    #         return
    #
    #     delete_scenario_from_excel_file(selected_indexes=selected_ids)
    #
    #     QMessageBox.information(self, "Success", "Selected scenarios deleted successfully.")
    #     self.close()  # Close the dialog after deletion

"""